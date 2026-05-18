from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from pack_video_text import clean_text, core_topic_text, hook_text, sentence_clip
from text_normalization import write_json_file


BOARD_HEADERS = [
    "优先级",
    "交接状态",
    "入选溯源",
    "视频 / 链接",
    "为什么值得研究",
    "适合复用在哪",
    "推荐深拆方向",
    "下一步场景",
    "核心主题",
    "钩子强度",
    "表现信号",
    "TikTok Shop 信号",
    "发布时间窗口",
    "商业置信度",
]


def build_board_row(
    video: dict,
    index: int,
    *,
    handoff_status: str,
    provenance: str,
    study_value: str,
    reuse_fit: str,
    teardown_direction: str,
    next_scene: str,
    metric_summary: str,
    publish_window: str,
) -> list[str]:
    return [
        clean_text(video.get("shortlist_priority")) or f"P{index}",
        handoff_status,
        provenance,
        clean_text(video.get("video_url")),
        study_value,
        reuse_fit,
        teardown_direction,
        next_scene,
        sentence_clip(core_topic_text(video), limit=72) or "主题文本缺失",
        sentence_clip(hook_text(video), limit=76) or "钩子文本缺失",
        metric_summary,
        clean_text(video.get("tkshop_signal")) or "未检测到",
        publish_window,
        str(video.get("commerce_confidence", "")),
    ]


def build_collection_board_payload(
    ranked_videos: list[dict],
    *,
    row_factory,
    aggregate_summary: dict | None = None,
    handoff_gate: str = "",
) -> dict:
    rows = []
    for index, video in enumerate(ranked_videos, start=1):
        rows.append(row_factory(video, index))
    return {
        "schema_version": "scene01-collection-board-v1",
        "headers": BOARD_HEADERS,
        "rows": rows,
        "row_count": len(rows),
        "handoff_gate": handoff_gate,
        "category": clean_text((aggregate_summary or {}).get("category")),
        "market": clean_text((aggregate_summary or {}).get("market")),
    }


def persist_collection_board_json(capture_root: Path, payload: dict) -> Path:
    output = capture_root / "collection_board.json"
    write_json_file(output, payload)
    return output


def write_feishu_style_board_xlsx(output_path: Path, payload: dict) -> Path:
    workbook = Workbook()
    ws = workbook.active
    ws.title = "采集看板"
    meta = [
        ("品类", payload.get("category") or "—"),
        ("市场", payload.get("market") or "—"),
        ("候选行数", str(payload.get("row_count", 0))),
        ("交接闸门", payload.get("handoff_gate") or "—"),
    ]
    ws["A1"] = "Scene 01 采集看板（Feishu 式主表）"
    for offset, (label, value) in enumerate(meta, start=2):
        ws.cell(row=offset, column=1, value=label)
        ws.cell(row=offset, column=2, value=value)
    header_row = len(meta) + 3
    headers = payload.get("headers") or BOARD_HEADERS
    for column_index, header in enumerate(headers, start=1):
        ws.cell(row=header_row, column=column_index, value=header)
    for row_offset, row in enumerate(payload.get("rows") or [], start=header_row + 1):
        for column_index, value in enumerate(row, start=1):
            ws.cell(row=row_offset, column=column_index, value=value)
    widths = [10, 14, 22, 28, 24, 20, 22, 14, 20, 22, 18, 16, 14, 12]
    for column_index, width in enumerate(widths[: len(headers)], start=1):
        ws.column_dimensions[get_column_letter(column_index)].width = width
    ws.freeze_panes = ws.cell(row=header_row + 1, column=1).coordinate
    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)
    return output_path


def persist_collection_board_exports(
    capture_root: Path,
    payload: dict,
) -> dict[str, str]:
    json_path = persist_collection_board_json(capture_root, payload)
    xlsx_path = write_feishu_style_board_xlsx(capture_root / "collection_board.xlsx", payload)
    return {"json": str(json_path), "xlsx": str(xlsx_path)}
